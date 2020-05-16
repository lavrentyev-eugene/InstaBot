from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import selenium.common.exceptions
from time import sleep
from datetime import datetime
from datetime import timedelta

import config  # Файл должен лежать в папке проекта

# Настройки бота из config.py
USERNAME = config.username
PASSWORD = config.password
SUB_PERIOD = config.sub_period
SUB_LIMIT = config.sub_limit
LIKES_LIMIT = config.likes_limit
POST_TIME = timedelta(minutes=config.post_time)
CHECK_SUB = config.CHECK_SUB

# Выбор браузера
browser = webdriver.Chrome()

# Открытие Excel-файла
wb = Workbook()
ws = wb.active
ws.title = USERNAME
ws['A1'] = 'Подписчики'
ws['B1'] = 'Подписки'
ws['C1'] = 'Невзаимные подписки'


# Ввод данных аккаунта
def registration():
    browser.get('https://www.instagram.com/')
    sleep(0.5)
    try:
        browser.find_element_by_xpath('//input[@name = "username"]').send_keys(USERNAME)
    except selenium.common.exceptions.NoSuchElementException:
        print('Не успел загрузиться сайт')
        sleep(2)
        browser.find_element_by_xpath('//input[@name = "username"]').send_keys(USERNAME)
    browser.find_element_by_xpath('//input[@name = "password"]').send_keys(PASSWORD)
    browser.find_element_by_xpath('//button[@type = "submit"]').click()
    sleep(3)


# Проверки всплывающих окон при входе в аккаунт
def login_check():
    # Проверка необходимости ввести код подтверждения
    try:
        browser.find_element_by_xpath('//button[contains(text(), "Отправить код безопасности")]').click()
        sleep(3)
        browser.find_element_by_xpath('//input[@type = "tel"]').send_keys(
            input('Введите код: '))
        sleep(1)
        browser.find_element_by_xpath('//button[contains(text(), "Отправить")]').click()
        browser.maximize_window()
        sleep(2)
    except selenium.common.exceptions.NoSuchElementException:
        pass

    # Проверка кнопок 'сохранить данные в браузере' и 'включить уведомления'
    try:
        browser.find_element_by_xpath('//button[contains(text(), "Не сейчас")]').click()
        print('Кнопка "Не сейчас" нажата')
        sleep(3)
    except selenium.common.exceptions.NoSuchElementException:
        print('Нет кнопки')
        sleep(3)
        browser.find_element_by_xpath('//button[contains(text(), "Не сейчас")]').click()
        print('Кнопка "Не сейчас" нажата')


# Осуществление прокрутки ленты до последнего элемента
def scroll(block):
    last_position, bottom = 0, 1
    while last_position != bottom:
        last_position = bottom
        sleep(1)
        bottom = browser.execute_script("""
            arguments[0].scrollTo(0 , arguments[0].scrollHeight)
            return arguments[0].scrollHeight
            """, block)


# проверка наличия подписки
def check_sub(block):
    sub = 0
    for _ in block.find_elements_by_xpath('//a[@title]'):
        while sub < SUB_LIMIT:
            try:
                browser.find_element_by_xpath('//button[contains(text(), "Подписаться")]').click()
                sub += 1
                sleep(SUB_PERIOD)
            except selenium.common.exceptions.NoSuchElementException:
                pass


# Получение списка всех подписчиков. Предусмотрена возможность взаимной подписки
def get_subscribers():
    browser.get(f'https://www.instagram.com/{USERNAME}')
    sleep(3)
    browser.find_element_by_xpath('//a[contains(@href,"/followers")]').click()
    sleep(2)
    subscribers_block = browser.find_element_by_xpath('/html/body/div[4]/div/div[2]')
    scroll(subscribers_block)
    if CHECK_SUB is True:
        check_sub(subscribers_block)
    global subscribers_links
    subscribers_links = [subscriber.get_attribute('href') for subscriber in
                         subscribers_block.find_elements_by_xpath('//a[@title]') if subscriber != '']
    for i in range(len(subscribers_links)):
        ws[f'B{i + 2}'] = subscribers_links[i]
        sleep(0.2)
        print(f'Внесено {i}/{len(subscribers_links)} человек')
    print('Подписчики внесены в базу данных')


# получение списка всех подписок
def get_followings():
    browser.get(f'https://www.instagram.com/{USERNAME}')
    sleep(2)
    browser.find_element_by_xpath('//a[contains(@href,"/following")]').click()
    sleep(2)
    followings_block = browser.find_element_by_xpath('//body/div[4]/div/div[2]')
    scroll(followings_block)
    global followings_links
    followings_links = [following.get_attribute('href') for following in
                        followings_block.find_elements_by_xpath('//a[@title]') if following != '']
    for i in range(len(followings_links)):
        ws[f'A{i + 2}'] = followings_links[i]
        sleep(0.2)
        print(f'Внесено {i}/{len(followings_links)} человек')
    print('Подписки внесены в базу данных')


# получение списка невзаимных подписок
def get_delta():
    delta = []
    for following in followings_links:
        if following in subscribers_links:
            delta.append(following)
    for i in range(len(delta)):
        ws[f'C{i + 2}'] = delta[i]
        sleep(0.2)
        print(f'Внесено {i}/{len(delta)} человек')
    print('Невзаимные подписки внесены в базу данных')


# Лайк постов с фильтрацией по времени появления
def like_posts():
    browser.get('https://www.instagram.com')
    sleep(2)
    posts_list = []
    likes = 0
    global article
    # Лайкает, пока не достигнут предел для количества лайков за одну сессию
    while likes < LIKES_LIMIT:
        try:
            if len(posts_list) >= 20:
                posts_list = posts_list[1:]
            sleep(1)
            browser.find_element_by_tag_name('body').send_keys(Keys.SPACE)
            posts = browser.find_elements_by_class_name('eo2As')  # Поиск всех постов на странице
            for post in posts:
                article = posts.index(post) + 1  # номер поста на странице
                # Попытка перейти к конкретному посту на странице и переход на предыдущий при неудаче
                try:
                    element = browser.find_element_by_xpath(f'//article[{article}]')
                except selenium.common.exceptions.NoSuchElementException:
                    element = browser.find_element_by_xpath(f'//article[{article - 1}]')

                # Фильтр просмотренных постов: проверка на фото или видео внутри
                try:
                    current_post = browser.find_element_by_xpath(
                        f'//article[{article}]/div[1]/div/div/div[1]/img').get_attribute('alt')
                except selenium.common.exceptions.NoSuchElementException:
                    current_post = browser.find_element_by_xpath(
                        f'//article[{article}]/div[1]/div/div/div[2]').get_attribute('class')

                if current_post not in posts_list:
                    print(current_post)
                    browser.execute_script('arguments[0].scrollIntoView(false)', element)  # Скролл к элементу
                    sleep(1)
                    posts_list.append(current_post)

                    # Проверка непролайканных постов на давность
                    try:
                        browser.find_element_by_xpath(
                            f'//article[{article}]/div[2]/section[1]/span[1]/\
                                                        button/*[name()="svg"][@aria-label="Нравится"]')
                        time_str = browser.find_element_by_xpath(
                            f'//article[{article}]/div[2]/div[2]/a/time').get_attribute('datetime')
                        time_format = '%Y-%m-%dT%H:%M:%S.%fZ'
                        time = datetime.strptime(time_str, time_format)
                        time_delta = datetime.utcnow() - time

                        # проверка на давность появления поста
                        if time_delta >= POST_TIME:
                            browser.find_element_by_xpath(
                                f'//article[{article}]/div[2]/section[1]/span[1]/\
                                button/*[name()="svg"][@aria-label="Нравится"]').click()
                            sleep(2)
                            likes += 1
                            print('Поставлен лайк')
                            sleep(SUB_PERIOD)
                    except selenium.common.exceptions.NoSuchElementException:
                        print('Уже стоит лайк')
                        sleep(1)
                        pass
                    except selenium.common.exceptions.ElementClickInterceptedException:
                        print('Неправильная прокрутка, элемент вне зоны видимости')
                        sleep(1)
                        pass
        # Обход ошибки, связанной с "брыканиями" сайта
        except selenium.common.exceptions.StaleElementReferenceException:
            sleep(5)
            pass


# Предупреждение о завершении работы бота и браузера
def quit_warning(warning=4):
    for second in reversed(range(1, warning)):
        print(f'Браузер закроется через {second}')
        sleep(1)
    browser.quit()  # Выход из браузера


registration()
login_check()
get_subscribers()
get_followings()
get_delta()
like_posts()
wb.save('subscribers.xlsx')  # Закрытие Excel-файла
quit_warning()
